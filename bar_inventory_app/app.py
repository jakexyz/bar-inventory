from flask import Flask, render_template, request, redirect, url_for, send_file, flash
from flask_sqlalchemy import SQLAlchemy
from datetime import datetime
from sqlalchemy import text
import math, csv, io, os, time
from collections import defaultdict

app = Flask(__name__)

# --- Secrets ---
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'change-me')

# --- Database URL (Render: set DATABASE_URL to your Postgres URL) ---
db_url = os.environ.get('DATABASE_URL', 'sqlite:///inventory.db')  # local fallback

# Normalize for SQLAlchemy/psycopg2
if db_url.startswith('postgres://'):
    db_url = db_url.replace('postgres://', 'postgresql+psycopg2://', 1)
elif db_url.startswith('postgresql://'):
    db_url = db_url.replace('postgresql://', 'postgresql+psycopg2://', 1)

# Enforce SSL on hosted Postgres
if db_url.startswith('postgresql+psycopg2://') and 'sslmode=' not in db_url:
    db_url += ('&' if '?' in db_url else '?') + 'sslmode=require'

app.config['SQLALCHEMY_DATABASE_URI'] = db_url
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# Resilient connection pool
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
    'pool_pre_ping': True,        # verify connection before using it
    'pool_recycle': 300,          # avoid stale idle connections
    'pool_size': 5,
    'max_overflow': 10,
    'connect_args': (
        {'sslmode': 'require'} if db_url.startswith('postgresql+psycopg2://') else {}
    ),
}

db = SQLAlchemy(app)

# --- Model ---
class Item(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(120), nullable=False)
    category = db.Column(db.String(80), default='Spirits')  # Vodka, Agave, etc.
    unit = db.Column(db.String(30), default='bottle')       # bottle, can, keg, ml
    case_size = db.Column(db.Integer, default=1)            # units per case
    par_cases = db.Column(db.Integer, nullable=True)        # par in cases
    par_units = db.Column(db.Integer, nullable=True)        # optional override in units
    current_units = db.Column(db.Integer, default=0)        # on-hand in units
    vendor = db.Column(db.String(120), nullable=True)
    cost_per_case = db.Column(db.Float, nullable=True)
    lead_time_days = db.Column(db.Integer, nullable=True)
    notes = db.Column(db.Text, nullable=True)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    def par_in_units(self):
        if self.par_units is not None:
            return self.par_units
        if self.par_cases is not None and self.case_size:
            return self.par_cases * self.case_size
        return None

    def needed_units(self):
        p = self.par_in_units()
        if p is None:
            return None
        return p - (self.current_units or 0)

    def cases_to_order(self):
        p = self.par_in_units()
        if p is None or not self.case_size or self.case_size <= 0:
            return 0
        needed = p - (self.current_units or 0)
        if needed <= 0:
            return 0
        return math.ceil(needed / self.case_size)

# --- Lazy DB init (avoid import-time create_all) ---
import threading

_init_lock = threading.Lock()
_db_inited = False

@app.before_request
def _ensure_db_initialized():
    global _db_inited
    if _db_inited:
        return
    with _init_lock:
        if _db_inited:
            return
        try:
            db.create_all()
            # Optional: gated seeding if you ever need it
            if os.environ.get('SEED_ON_START') == '1':
                try:
                    # put your seeding here (e.g., import CSV if tables empty)
                    pass
                except Exception:
                    app.logger.exception('Auto-seed failed')
        except Exception:
            app.logger.exception('DB init skipped due to error')
        _db_inited = True


# ---------- Views ----------
@app.route('/')
def index():
    q = request.args.get('q', '').strip().lower()
    cat = request.args.get('category', '').strip()
    vendor = request.args.get('vendor', '').strip()
    only_to_order = request.args.get('to_order') == '1'
    group = request.args.get('group', '')  # '', 'vendor'

    items = Item.query
    if q:
        from sqlalchemy import or_
        items = items.filter(or_(
            Item.name.ilike(f"%{q}%"),
            Item.category.ilike(f"%{q}%"),
            Item.vendor.ilike(f"%{q}%"),
        ))
    if cat:
        items = items.filter(Item.category == cat)
    if vendor:
        items = items.filter(Item.vendor == vendor)

    items = items.order_by(Item.vendor.asc(), Item.category.asc(), Item.name.asc()).all()
    if only_to_order:
        items = [i for i in items if i.cases_to_order() > 0]

    categories = sorted(set([i.category for i in Item.query.all()]))
    vendors = sorted(set([i.vendor for i in Item.query.all() if i.vendor]))

    grouped = {}
    if group == 'vendor':
        for i in items:
            vkey = i.vendor or 'Unassigned Vendor'
            grouped.setdefault(vkey, {})
            grouped[vkey].setdefault(i.category or 'Other', []).append(i)

    return render_template('index.html',
                           items=items, q=q,
                           categories=categories, vendors=vendors,
                           selected_category=cat, selected_vendor=vendor,
                           only_to_order=only_to_order, group=group, grouped=grouped)

# CRUD
@app.route('/item/new', methods=['GET', 'POST'])
def new_item():
    if request.method == 'POST':
        d = request.form
        item = Item(
            name=d.get('name').strip(),
            category=d.get('category') or 'Spirits',
            unit=d.get('unit') or 'bottle',
            case_size=int(d.get('case_size') or 0),
            par_cases=int(d.get('par_cases') or 0) if d.get('par_cases') else None,
            par_units=int(d.get('par_units') or 0) if d.get('par_units') else None,
            current_units=int(d.get('current_units') or 0),
            vendor=d.get('vendor') or None,
            cost_per_case=float(d.get('cost_per_case')) if d.get('cost_per_case') else None,
            lead_time_days=int(d.get('lead_time_days')) if d.get('lead_time_days') else None,
            notes=d.get('notes') or None,
        )
        db.session.add(item); db.session.commit()
        flash('Item created.', 'success')
        return redirect(url_for('index'))
    return render_template('edit.html', item=None)

@app.route('/item/<int:item_id>/edit', methods=['GET','POST'])
def edit_item(item_id):
    item = Item.query.get_or_404(item_id)
    if request.method == 'POST':
        d = request.form
        item.name = d.get('name').strip()
        item.category = d.get('category') or 'Spirits'
        item.unit = d.get('unit') or 'bottle'
        item.case_size = int(d.get('case_size') or 0)
        item.par_cases = int(d.get('par_cases') or 0) if d.get('par_cases') else None
        item.par_units = int(d.get('par_units') or 0) if d.get('par_units') else None
        item.current_units = int(d.get('current_units') or 0)
        item.vendor = d.get('vendor') or None
        item.cost_per_case = float(d.get('cost_per_case')) if d.get('cost_per_case') else None
        item.lead_time_days = int(d.get('lead_time_days')) if d.get('lead_time_days') else None
        item.notes = d.get('notes') or None
        db.session.commit()
        flash('Item updated.', 'success')
        return redirect(url_for('index'))
    return render_template('edit.html', item=item)

@app.route('/item/<int:item_id>/delete', methods=['POST'])
def delete_item(item_id):
    item = Item.query.get_or_404(item_id)
    db.session.delete(item); db.session.commit()
    flash('Item deleted.', 'info')
    return redirect(url_for('index'))

# Import / Export CSV
@app.route('/export')
def export_csv():
    out = io.StringIO()
    w = csv.writer(out)
    w.writerow(['name','category','unit','case_size','par_cases','par_units','current_units','vendor','cost_per_case','lead_time_days','notes'])
    for i in Item.query.order_by(Item.category.asc(), Item.name.asc()).all():
        w.writerow([i.name, i.category, i.unit, i.case_size, i.par_cases or '', i.par_units or '', i.current_units or 0, i.vendor or '', i.cost_per_case or '', i.lead_time_days or '', i.notes or ''])
    out.seek(0)
    return send_file(io.BytesIO(out.getvalue().encode('utf-8')), mimetype='text/csv', as_attachment=True, download_name='bar_inventory_export.csv')

@app.route('/import', methods=['GET','POST'])
def import_csv():
    if request.method == 'POST':
        f = request.files.get('file')
        if not f:
            flash('No file selected.', 'warning'); return redirect(url_for('import_csv'))
        stream = io.StringIO(f.stream.read().decode('utf-8'))
        reader = csv.DictReader(stream)
        new_count = 0
        for row in reader:
            try:
                name_val = (row.get('name') or '').strip()
                if not name_val: 
                    continue
                category_val = (row.get('category') or 'Spirits').strip() or 'Spirits'
                vendor_val = (row.get('vendor') or None)
                existing = db.session.query(Item).filter(
                    db.func.lower(Item.name)==name_val.lower(),
                    db.func.lower(db.func.coalesce(Item.vendor,''))==(vendor_val or '').lower(),
                    db.func.lower(db.func.coalesce(Item.category,''))==category_val.lower()
                ).first()
                if existing:
                    if (row.get('case_size') or '').strip() != '': existing.case_size = int(row.get('case_size'))
                    if (row.get('par_cases') or '').strip() != '': existing.par_cases = int(row.get('par_cases'))
                    if (row.get('par_units') or '').strip() != '': existing.par_units = int(row.get('par_units'))
                    cur = int(row.get('current_units') or 0)
                    existing.current_units = max(existing.current_units or 0, cur)
                    if (row.get('cost_per_case') or '').strip() != '': existing.cost_per_case = float(row.get('cost_per_case'))
                    if (row.get('lead_time_days') or '').strip() != '': existing.lead_time_days = int(row.get('lead_time_days'))
                    if row.get('notes'): existing.notes = (existing.notes + ' | ' if existing.notes else '') + row['notes']
                else:
                    item = Item(
                        name=name_val,
                        category=category_val,
                        unit=(row.get('unit') or 'bottle').strip() or 'bottle',
                        case_size=int(row.get('case_size') or 0) if (row.get('case_size') or '').strip() != '' else 0,
                        par_cases=int(row.get('par_cases')) if (row.get('par_cases') or '').strip() != '' else None,
                        par_units=int(row.get('par_units')) if (row.get('par_units') or '').strip() != '' else None,
                        current_units=int(row.get('current_units') or 0),
                        vendor=vendor_val,
                        cost_per_case=float(row.get('cost_per_case')) if (row.get('cost_per_case') or '').strip() != '' else None,
                        lead_time_days=int(row.get('lead_time_days')) if (row.get('lead_time_days') or '').strip() != '' else None,
                        notes=row.get('notes') or None,
                    )
                    db.session.add(item); new_count += 1
            except Exception as e:
                print('Row error:', e, row)
        db.session.commit()
        flash(f'Imported {new_count} new items (existing updated).', 'success')
        return redirect(url_for('index'))
    return render_template('import.html')

# De-dupe helper
@app.route('/admin/dedupe')
def admin_dedupe():
    items = Item.query.order_by(Item.id.asc()).all()
    def _norm(s): return (s or '').strip().lower()
    groups = {}
    for it in items:
        k = (_norm(it.vendor), _norm(it.category), _norm(it.name))
        groups.setdefault(k, []).append(it)
    removed = 0
    for k, lst in groups.items():
        if len(lst) <= 1: continue
        keep = lst[0]
        for dup in lst[1:]:
            keep.current_units = max(keep.current_units or 0, dup.current_units or 0)
            if not keep.case_size and dup.case_size: keep.case_size = dup.case_size
            if keep.par_cases is None and dup.par_cases is not None: keep.par_cases = dup.par_cases
            if keep.par_units is None and dup.par_units is not None: keep.par_units = dup.par_units
            if keep.cost_per_case is None and dup.cost_per_case is not None: keep.cost_per_case = dup.cost_per_case
            if keep.lead_time_days is None and dup.lead_time_days is not None: keep.lead_time_days = dup.lead_time_days
            if dup.notes: keep.notes = (keep.notes + ' | ' if keep.notes else '') + dup.notes
            db.session.delete(dup); removed += 1
    db.session.commit()
    flash(f'Removed {removed} duplicates.', 'success')
    return redirect(url_for('index'))

# --- Readiness & Health ---
@app.route('/admin/ready')
def admin_ready():
    """DB-free readiness endpoint for Render health checks."""
    return {'status': 'ok'}, 200

@app.route('/admin/health')
def admin_health():
    """
    DB-connected health endpoint.
    Returns 200 if DB is reachable and can run SELECT 1; 503 otherwise.
    """
    started = time.time()
    try:
        # Use a fresh connection to verify real connectivity (not a stale pooled handle)
        with db.engine.connect() as conn:
            # Optional: keep it very fast; ignore if unsupported
            try:
                conn.execute(text("SET LOCAL statement_timeout = 2000"))
            except Exception:
                pass
            conn.execute(text("SELECT 1"))
        duration_ms = int((time.time() - started) * 1000)
        return {
            'status': 'ok',
            'db': 'connected',
            'duration_ms': duration_ms
        }, 200
    except Exception as e:
        app.logger.exception('DB health check failed')
        duration_ms = int((time.time() - started) * 1000)
        return {
            'status': 'degraded',
            'db': 'error',
            'duration_ms': duration_ms,
            'error': str(e)
        }, 503

# Optional: keep the old template-based metrics as a separate page
@app.route('/admin/db-metrics')
def admin_db_metrics():
    total = Item.query.count()
    missing_case = Item.query.filter((Item.case_size == None) | (Item.case_size == 0)).count()
    missing_par_cases = Item.query.filter(Item.par_cases == None).count()
    return render_template('health.html', total=total, missing_case=missing_case, missing_par_cases=missing_par_cases)

# --- What to Order (on-page) ---
@app.route('/order')
def order_summary():
    """Show what to order today (no Excel needed)."""
    q = request.args.get('q', '').strip().lower()
    cat = request.args.get('category', '').strip()
    vendor = request.args.get('vendor', '').strip()

    items = Item.query
    if q:
        from sqlalchemy import or_
        items = items.filter(or_(
            Item.name.ilike(f"%{q}%"),
            Item.category.ilike(f"%{q}%"),
            Item.vendor.ilike(f"%{q}%"),
        ))
    if cat:
        items = items.filter(Item.category == cat)
    if vendor:
        items = items.filter(Item.vendor == vendor)
    items = items.order_by(Item.vendor.asc(), Item.category.asc(), Item.name.asc()).all()

    def nz(val, default=0):
        try:
            if val is None: return default
            return int(val)
        except Exception:
            try:
                return int(float(val))
            except Exception:
                return default

    rows = []
    vendors = defaultdict(lambda: defaultdict(list))
    vendor_totals = {}
    grand_total = 0.0

    for it in items:
        case_size = nz(it.case_size, 0)
        par_units = nz(it.par_in_units(), 0)
        on_hand = nz(it.current_units, 0)
        need_units = max(par_units - on_hand, 0)
        if need_units <= 0:
            continue
        order_cases = (need_units + case_size - 1) // case_size if case_size else 0
        est_total = (order_cases * (it.cost_per_case or 0)) if (order_cases and it.cost_per_case) else 0
        row = dict(
            vendor=it.vendor or "Unassigned Vendor",
            category=it.category or "",
            name=it.name,
            case_size=case_size,
            par_units=par_units,
            on_hand=on_hand,
            need_units=need_units,
            order_cases=order_cases,
            cost_per_case=it.cost_per_case,
            est_total=est_total,
            notes=it.notes or ""
        )
        vendors[row["vendor"]][row["category"]].append(row)
        vendor_totals.setdefault(row["vendor"], 0.0)
        vendor_totals[row["vendor"]] += est_total
        grand_total += est_total

    return render_template('order.html',
                           vendors=vendors, vendor_totals=vendor_totals,
                           grand_total=grand_total, count=sum(len(x) for cat in vendors.values() for x in cat.values()))

# --- Excel export (optional) ---
@app.route('/order.xlsx')
def order_excel():
    """Download Excel. Default: only items to order. Add ?all=1 for ALL items."""
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    import datetime as _dt

    q = request.args.get('q', '').strip().lower()
    cat = request.args.get('category', '').strip()
    vendor = request.args.get('vendor', '').strip()
    include_all = request.args.get('all') == '1'

    items = Item.query
    if q:
        from sqlalchemy import or_
        items = items.filter(or_(
            Item.name.ilike(f"%{q}%"),
            Item.category.ilike(f"%{q}%"),
            Item.vendor.ilike(f"%{q}%"),
        ))
    if cat:
        items = items.filter(Item.category == cat)
    if vendor:
        items = items.filter(Item.vendor == vendor)
    items = items.order_by(Item.vendor.asc(), Item.category.asc(), Item.name.asc()).all()

    def nz(val, default=0):
        try:
            if val is None: return default
            return int(val)
        except Exception:
            try:
                return int(float(val))
            except Exception:
                return default

    rows = []
    for it in items:
        case_size = nz(it.case_size, 0)
        par_units = nz(it.par_in_units(), 0)
        on_hand = nz(it.current_units, 0)
        need_units = max(par_units - on_hand, 0)
        order_cases = 0
        if case_size and need_units > 0:
            order_cases = (need_units + case_size - 1) // case_size
        rows.append({
            "vendor": it.vendor or "Unassigned Vendor",
            "category": it.category or "",
            "name": it.name,
            "case_size": case_size,
            "par_units": par_units,
            "on_hand": on_hand,
            "need_units": need_units,
            "order_cases": order_cases,
            "cost_per_case": it.cost_per_case or "",
            "est_total": (order_cases * (it.cost_per_case or 0)) if (order_cases and it.cost_per_case) else "",
            "notes": it.notes or ""
        })

    data = rows if include_all else [r for r in rows if r["order_cases"] and r["order_cases"] > 0]

    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    by_vendor = defaultdict(list)
    for r in data:
        by_vendor[r["vendor"]].append(r)

    grand_total = 0.0
    if not by_vendor:
        ws = wb.create_sheet("No Items")
        ws.append(["Nothing to export based on current inputs/filters."])
    else:
        for vname, vrows in by_vendor.items():
            ws = wb.create_sheet(title=(vname[:31] if vname else "Vendor"))
            headers = ["Vendor","Category","Item","Case Size","Par (units)","On Hand (units)","Need (units)","Order (cases)","Cost/Case","Est. Total","Notes"]
            ws.append(headers)
            vendor_total = 0.0
            for r in vrows:
                est = r["est_total"] if isinstance(r["est_total"], (int, float)) else 0
                vendor_total += est
                ws.append([
                    vname, r["category"], r["name"], r["case_size"],
                    r["par_units"], r["on_hand"], r["need_units"], r["order_cases"],
                    r["cost_per_case"], r["est_total"], r["notes"]
                ])
            for col_idx in range(1, len(headers)+1):
                max_len = 0
                for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=True):
                    val = row[0]
                    if val is None: continue
                    max_len = max(max_len, len(str(val)))
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)
            ws.append([])
            ws.append(["", "", "", "", "", "", "", "Vendor Total", "", vendor_total, ""])
            grand_total += vendor_total

    sumws = wb.create_sheet("All Vendors")
    sumws.append(["Generated", _dt.date.today().isoformat()])
    sumws.append(["Grand Total (if costs set)", grand_total])

    fname = f"bar_order_{_dt.date.today().isoformat()}{'_all' if include_all else ''}.xlsx"
    bio = io.BytesIO()
    wb.save(bio); bio.seek(0)
    return send_file(bio, as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
